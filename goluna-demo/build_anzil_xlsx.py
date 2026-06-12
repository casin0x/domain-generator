#!/usr/bin/env python3
"""Build Anzil_Brazil_Pilot_Forecast.xlsx — Sheet 1: Path C cost forecast, Sheet 2: revenue targets."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

ARIAL = "Arial"
BLUE = Font(name=ARIAL, size=10, color="0000FF")          # hardcoded inputs
BLACK = Font(name=ARIAL, size=10)                          # formulas
GREEN = Font(name=ARIAL, size=10, color="008000")          # cross-sheet links
BOLD = Font(name=ARIAL, size=10, bold=True)
H1 = Font(name=ARIAL, size=14, bold=True)
SUB = Font(name=ARIAL, size=9, italic=True, color="666666")
HDR_FILL = PatternFill("solid", start_color="1F1F3A")
HDR_FONT = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
ANZIL_FILL = PatternFill("solid", start_color="EDE9FE")
YELLOW = PatternFill("solid", start_color="FFFF00")
TOT_FILL = PatternFill("solid", start_color="EFEFEF")
CUR = '$#,##0;[Red]($#,##0);"-"'
PROFIT = '+$#,##0;($#,##0);"-"'
PCT = "0.0%"
NUM = '#,##0;[Red](#,##0);"-"'
GOOD_RULE = lambda: CellIsRule(operator="greaterThan", formula=["0"],
    fill=PatternFill("solid", start_color="C6EFCE"), font=Font(name=ARIAL, size=10, bold=True, color="006100"))
BAD_RULE = lambda: CellIsRule(operator="lessThan", formula=["0"],
    fill=PatternFill("solid", start_color="FFC7CE"), font=Font(name=ARIAL, size=10, bold=True, color="9C0006"))
def profit_color(ws, rng):
    ws.conditional_formatting.add(rng, GOOD_RULE())
    ws.conditional_formatting.add(rng, BAD_RULE())
thin = Side(style="thin", color="CCCCCC")
BORDER = Border(top=thin, bottom=thin, left=thin, right=thin)

wb = Workbook()

# ---------------- Sheet 1: Cost Forecast ----------------
ws = wb.active
ws.title = "Cost Forecast"

ws["A1"] = "GoLuna — Path C (Anzil) Month-by-Month Cost Forecast (USD)"
ws["A1"].font = H1
ws.merge_cells("A1:N1")
ws["A2"] = ("Anzil engagement hard-capped at $100K over the 3-month pilot (M2–M4) · 1M SEK dedicated to Brazil/Anzil · "
            "influencer activation removed from Anzil scope · Clipping $25K (M2 prepay) runs via a separate vendor outside the cap · "
            "M2 kill checkpoint. Benchmarks, not GoLuna-measured data — verify at M4.")
ws["A2"].font = SUB
ws.merge_cells("A2:N2")

# assumptions
ws["A4"] = "Assumptions"; ws["A4"].font = BOLD
rows_a = [
    ("FX (SEK per USD)", 10.5, BLUE, "0.0", True),
    ("Working budget (2.5M SEK, USD)", "=2500000/B5", BLACK, CUR, False),
    ("Brazil/Anzil envelope (1M SEK, USD)", "=1000000/B5", BLACK, CUR, False),
    ("Anzil management fee (% of media)", 0.15, BLUE, PCT, True),
    ("Anzil contract hard cap (USD)", 100000, BLUE, CUR, True),
]
for i, (label, val, font, fmt, yellow) in enumerate(rows_a, start=5):
    ws.cell(row=i, column=1, value=label).font = BLACK
    c = ws.cell(row=i, column=2, value=val); c.font = font; c.number_format = fmt
    if yellow: c.fill = YELLOW

HDR = 11
ws.cell(row=HDR, column=1, value="Line")
for m in range(1, 13):
    ws.cell(row=HDR, column=1 + m, value=f"M{m}")
ws.cell(row=HDR, column=14, value="Total")
for col in range(1, 15):
    c = ws.cell(row=HDR, column=col); c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = Alignment(horizontal="right" if col > 1 else "left")

# (label, [12 monthly values], is_anzil)
lines = [
    ("Salaries (T+O+G)", [5000]*12, False),
    ("X VIP Transfer Method", [0]+[1143]*11, False),
    ("Cubeia AWS hosting", [1836]*12, False),
    ("Customer.io (Cubeia App. B)", [324]*12, False),
    ("Freshchat (Cubeia App. B)", [0]+[54]*11, False),
    ("Cubeia platform min (from M7)", [0]*6+[5400]*6, False),
    ("Security stack subs", [281]*12, False),
    ("Claude Max", [200]*12, False),
    ("Security stack — hardware (one-time)", [1652]+[0]*11, False),
    ("Costa Rica corp (2026)", [6858]+[0]*11, False),
    ("Designer (casino design)", [4800]+[0]*11, False),
    ("Betby setup (sportsbook)", [3000]+[0]*11, False),
    ("Other subscriptions", [476]*12, False),
    ("Clipping (content/streamer campaign · prepaid M2 · separate vendor, not Anzil)", [0, 25000]+[0]*10, True),
    ("Anzil setup (€3K, one-time, M1)", [3240]+[0]*11, True),
    ("Anzil retainer (€1.5K/mo, M2–M4)", [0]+[1620]*3+[0]*8, True),
    ("Anzil media — Meta+WhatsApp (M2–M4)", [0, 15000, 25000, 29000]+[0]*8, True),
    ("Anzil management (15% of media)", None, True),  # formula row
]
FIRST = HDR + 1
CLIP_ROW = FIRST + 13
MEDIA_ROW = FIRST + 16
for i, (label, vals, is_anzil) in enumerate(lines):
    r = FIRST + i
    lc = ws.cell(row=r, column=1, value=label); lc.font = BOLD if is_anzil else BLACK
    for m in range(12):
        col = 2 + m
        if vals is None:  # management = 15% x media
            c = ws.cell(row=r, column=col, value=f"={get_column_letter(col)}{MEDIA_ROW}*$B$8")
            c.font = BLACK
        else:
            c = ws.cell(row=r, column=col, value=vals[m]); c.font = BLUE
        c.number_format = CUR
        if is_anzil: c.fill = ANZIL_FILL
    t = ws.cell(row=r, column=14, value=f"=SUM(B{r}:M{r})"); t.font = BLACK; t.number_format = CUR
    if is_anzil: t.fill = ANZIL_FILL
LAST = FIRST + len(lines) - 1
ANZIL_FIRST, ANZIL_LAST = FIRST + 14, FIRST + 17  # setup..management
FIX_LAST = FIRST + 12                              # GoLuna fixed block (clipping excluded)

# subtotal rows: marketing (Anzil + clipping) and GoLuna fixed, per month
MKT = LAST + 1
ws.cell(row=MKT, column=1, value="MARKETING — monthly total (Anzil + clipping)").font = BOLD
for m in range(13):
    col = get_column_letter(2 + m)
    rng = f"{col}{ANZIL_FIRST}:{col}{ANZIL_LAST}"
    c = ws.cell(row=MKT, column=2 + m, value=f"=SUM({rng})+{col}{CLIP_ROW}")
    c.font = BOLD; c.number_format = CUR; c.fill = ANZIL_FILL
FIX = MKT + 1
ws.cell(row=FIX, column=1, value="GoLuna fixed — monthly total (salaries, platform, security, one-times)").font = BOLD
for m in range(13):
    col = get_column_letter(2 + m)
    c = ws.cell(row=FIX, column=2 + m, value=f"=SUM({col}{FIRST}:{col}{FIX_LAST})")
    c.font = BOLD; c.number_format = CUR; c.fill = TOT_FILL

TOT = FIX + 1
ws.cell(row=TOT, column=1, value="Monthly total (fixed + marketing)").font = BOLD
for m in range(13):
    col = get_column_letter(2 + m)
    c = ws.cell(row=TOT, column=2 + m, value=f"={col}{MKT}+{col}{FIX}")
    c.font = BOLD; c.number_format = CUR; c.fill = TOT_FILL

CUM = TOT + 1
ws.cell(row=CUM, column=1, value="Cumulative").font = BOLD
ws.cell(row=CUM, column=2, value=f"=B{TOT}").number_format = CUR
for m in range(1, 12):
    col, prev = get_column_letter(2 + m), get_column_letter(1 + m)
    c = ws.cell(row=CUM, column=2 + m, value=f"={prev}{CUM}+{col}{TOT}")
    c.font = BLACK; c.number_format = CUR
ws.cell(row=CUM, column=2).font = BLACK

REM = CUM + 1
ws.cell(row=REM, column=1, value="Operating budget remaining (vs 2.5M SEK)").font = BOLD
for m in range(12):
    col = get_column_letter(2 + m)
    c = ws.cell(row=REM, column=2 + m, value=f"=$B$6-{col}{CUM}")
    c.font = BLACK; c.number_format = CUR

# Anzil summary block
S = REM + 3
ws.cell(row=S, column=1, value="Anzil engagement summary").font = BOLD
items = [
    ("Anzil committed total (setup + retainer + media + mgmt)", f"=N{ANZIL_FIRST}+N{ANZIL_FIRST+1}+N{ANZIL_FIRST+2}+N{ANZIL_FIRST+3}", CUR, BLACK),
    ("Brazil/Anzil envelope (1M SEK)", "=B7", CUR, BLACK),
    ("Headroom vs envelope (WhatsApp/PIX fees, TBD)", f"=B7-B{S+1}", CUR, BLACK),
    ("Contract hard cap", "=B9", CUR, BLACK),
    ("Headroom vs hard cap", f"=B9-B{S+1}", CUR, BLACK),
    ("Sunk if cut at M2 checkpoint (setup + M2 retainer/media/mgmt)", f"=B{ANZIL_FIRST}+C{ANZIL_FIRST+1}+C{ANZIL_FIRST+2}+C{ANZIL_FIRST+3}", CUR, BLACK),
    ("Never deployed if cut at M2", f"=B{S+1}-B{S+6}", CUR, BLACK),
    ("Paid if cut after 2 live months (end of M3)", f"=B{S+6}+D{ANZIL_FIRST+1}+D{ANZIL_FIRST+2}+D{ANZIL_FIRST+3}", CUR, BLACK),
    ("Clipping campaign (separate vendor — outside the Anzil cap and 1M SEK envelope; funded from the experiments pot)", f"=N{CLIP_ROW}", CUR, BLACK),
]
for i, (label, formula, fmt, font) in enumerate(items, start=S + 1):
    ws.cell(row=i, column=1, value=label).font = BLACK
    c = ws.cell(row=i, column=2, value=formula); c.font = font; c.number_format = fmt

# what is deliberately NOT in this sheet
NB = S + len(items) + 2
ws.cell(row=NB, column=1, value="Not in this sheet (and why)").font = BOLD
notes = [
    "Player bonuses & incentives (welcome / reload / cashback ≈ 25% of GGR) — variable cost, scales with revenue: deducted on the Revenue Targets sheet (GGR → NGR → net contribution).",
    "Game-provider / Cubeia 5% / crypto / PIX fees — also % of GGR, inside the 51.5% net-contribution margin on the Revenue Targets sheet.",
    "Promotional pool (community giveaways — $4K/mo M2–M4 in Paths A/B) — NOT budgeted for Path C; add ~$12K if the Anzil pilot should run community incentives too.",
    "WhatsApp per-message + Anzil payments fees — volume-tiered, TBD at integration; covered by the cap headroom.",
    "Bankroll 500K SEK (~$48K) — held separately for player withdrawals, not an operating cost.",
]
for i, t in enumerate(notes, start=NB + 1):
    ws.cell(row=i, column=1, value="• " + t).font = SUB
    ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=14)

ws.column_dimensions["A"].width = 44
for col in range(2, 15):
    ws.column_dimensions[get_column_letter(col)].width = 11
ws.freeze_panes = "B12"

# ---------------- Sheet 2: Revenue Targets ----------------
w2 = wb.create_sheet("Revenue Targets")
w2["A1"] = "GoLuna — Path C Revenue Targets by Month (what to optimize for)"
w2["A1"].font = H1
w2.merge_cells("A1:H1")
w2["A2"] = ("NEED = ~28% retention (the break-even line: M4-exit run-rate covers the M7+ fixed burn). "
            "OPTIMIZE = ~39% retention (the flywheel compounds). Basis: ARPU $30/mo (Brazil regulated avg), "
            "CAC $20 (Anzil quote), net contribution 51.5% of GGR (incl. ~5% high-estimate PIX fee). "
            "Market benchmarks, not GoLuna data — recalibrate with real cohorts at M2–M3.")
w2["A2"].font = SUB
w2.merge_cells("A2:H2")

w2["A4"] = "Assumptions"; w2["A4"].font = BOLD
rows_b = [
    ("ARPU (GGR / active / month)", 30, BLUE, CUR, True),
    ("CAC per FTD (Anzil quote $12–25)", 20, BLUE, CUR, True),
    ("NGR (GGR − bonuses 25%)", 0.75, BLUE, PCT, False),
    ("Net contribution (after all variable, incl. PIX high-est.)", 0.515, BLUE, PCT, True),
    ("Retention — NEED (break-even line)", 0.28, BLUE, PCT, True),
    ("Retention — OPTIMIZE (compounding)", 0.39, BLUE, PCT, True),
    ("M7+ fixed run-rate (from Cost Forecast)", f"='Cost Forecast'!H{TOT}", GREEN, CUR, False),
]
for i, (label, val, font, fmt, yellow) in enumerate(rows_b, start=5):
    w2.cell(row=i, column=1, value=label).font = BLACK
    c = w2.cell(row=i, column=2, value=val); c.font = font; c.number_format = fmt
    if yellow: c.fill = YELLOW

T = 20  # monthly-detail header row; rows 13-18 hold the PROFIT AT A GLANCE block (filled in after the detail blocks exist)
heads = ["Metric", "M1 (onboard)", "M2 (pilot 1)", "M3 (pilot 2)", "M4 (pilot 3 · gate)", "M5", "M6", "M7+ /mo"]
for j, h in enumerate(heads):
    c = w2.cell(row=T, column=1 + j, value=h); c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = Alignment(horizontal="right" if j else "left")

mr = MEDIA_ROW
r_media, r_ftd, r_cum = T + 1, T + 2, T + 3
w2.cell(row=r_media, column=1, value="Anzil media spend").font = BLACK
for j, src in enumerate(["B", "C", "D", "E", "F", "G", "H"]):
    c = w2.cell(row=r_media, column=2 + j, value=f"='Cost Forecast'!{src}{mr}")
    c.font = GREEN; c.number_format = CUR
w2.cell(row=r_ftd, column=1, value="New FTDs (media ÷ CAC)").font = BLACK
for j in range(7):
    col = get_column_letter(2 + j)
    c = w2.cell(row=r_ftd, column=2 + j, value=f"={col}{r_media}/$B$6")
    c.font = BLACK; c.number_format = NUM
w2.cell(row=r_cum, column=1, value="Cumulative FTDs (end of month)").font = BLACK
w2.cell(row=r_cum, column=2, value=f"=B{r_ftd}").number_format = NUM
for j in range(1, 7):
    col, prev = get_column_letter(2 + j), get_column_letter(1 + j)
    c = w2.cell(row=r_cum, column=2 + j, value=f"={prev}{r_cum}+{col}{r_ftd}")
    c.font = BLACK; c.number_format = NUM
w2.cell(row=r_cum, column=2).font = BLACK

def block(start, ret_cell, tag):
    r_act, r_ggr, r_bon, r_ngr, r_net, r_cost, r_res, r_cum2 = range(start, start + 8)
    w2.cell(row=start - 1, column=1, value=tag).font = BOLD
    w2.cell(row=start - 1, column=1).fill = TOT_FILL
    for j in range(7):
        w2.cell(row=start - 1, column=2 + j).fill = TOT_FILL
    w2.cell(row=r_act, column=1, value="Active depositors").font = BLACK
    for j in range(7):
        col = get_column_letter(2 + j)
        if j == 0:
            f = "=0"
        elif j <= 3:  # M2-M4: new + ret x cum prior
            prev = get_column_letter(1 + j)
            f = f"={col}{r_ftd}+{ret_cell}*{prev}{r_cum}"
        else:  # M5+: ret x total pilot FTDs (cum at M4 = col E)
            f = f"={ret_cell}*$E${r_cum}"
        c = w2.cell(row=r_act, column=2 + j, value=f); c.font = BLACK; c.number_format = NUM
    for r, label, mult in [
        (r_ggr, "Target GGR (actives × ARPU)", f"{r_act}*$B$5"),
        (r_bon, "− Player bonuses & incentives (25% of GGR)", f"{r_ggr}*($B$7-1)"),
        (r_ngr, "NGR  (= GGR − bonuses)", f"{r_ggr}*$B$7"),
        (r_net, "Net contribution (51.5% of GGR, after all variable fees)", f"{r_ggr}*$B$8"),
    ]:
        w2.cell(row=r, column=1, value=label).font = BLACK
        for j in range(7):
            col = get_column_letter(2 + j)
            c = w2.cell(row=r, column=2 + j, value=f"={col}{mult}")
            c.font = BLACK; c.number_format = CUR
    w2.cell(row=r_cost, column=1, value="Monthly cost (from Cost Forecast)").font = BLACK
    for j, src in enumerate(["B", "C", "D", "E", "F", "G", "H"]):
        c = w2.cell(row=r_cost, column=2 + j, value=f"='Cost Forecast'!{src}{TOT}")
        c.font = GREEN; c.number_format = CUR
    w2.cell(row=r_res, column=1, value="MONTHLY PROFIT / LOSS  (net contribution − cost)").font = BOLD
    for j in range(7):
        col = get_column_letter(2 + j)
        c = w2.cell(row=r_res, column=2 + j, value=f"={col}{r_net}-{col}{r_cost}")
        c.font = BOLD; c.number_format = PROFIT
    w2.cell(row=r_cum2, column=1, value="CUMULATIVE PROFIT / LOSS  (M1–M6)").font = BOLD
    w2.cell(row=r_cum2, column=2, value=f"=B{r_res}").number_format = PROFIT
    for j in range(1, 6):
        col, prev = get_column_letter(2 + j), get_column_letter(1 + j)
        c = w2.cell(row=r_cum2, column=2 + j, value=f"={prev}{r_cum2}+{col}{r_res}")
        c.font = BOLD; c.number_format = PROFIT
    w2.cell(row=r_cum2, column=8, value="n/a").font = SUB
    profit_color(w2, f"B{r_res}:H{r_res}")
    profit_color(w2, f"B{r_cum2}:G{r_cum2}")
    return r_res, r_cum2

need_res, need_cum = block(T + 6, "$B$9", "NEED — break-even line (~28% retention)")
opt_res, opt_cum = block(need_cum + 3, "$B$10", "OPTIMIZE — compounding (~39% retention)")
end_opt = opt_cum

# ---- PROFIT AT A GLANCE (rows 13-18, references the detail blocks below) ----
G = 13
w2.cell(row=G, column=1, value="PROFIT — AT A GLANCE").font = BOLD
w2.cell(row=G, column=1).fill = TOT_FILL
for j in range(1, 5): w2.cell(row=G, column=1 + j).fill = TOT_FILL
gheads = ["Scenario", "Pilot M1–M4 (the investment)", "M5–M6 / mo", "M7+ / mo", "Cumulative at M6"]
for j, h in enumerate(gheads):
    c = w2.cell(row=G + 1, column=1 + j, value=h); c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = Alignment(horizontal="right" if j else "left")
for i, (nm, r_res_x, r_cum_x) in enumerate([("NEED — break-even (~28% ret.)", need_res, need_cum),
                                            ("OPTIMIZE — compounding (~39% ret.)", opt_res, opt_cum)]):
    r = G + 2 + i
    w2.cell(row=r, column=1, value=nm).font = BOLD
    for col_i, f in [(2, f"=SUM(B{r_res_x}:E{r_res_x})"), (3, f"=F{r_res_x}"), (4, f"=H{r_res_x}"), (5, f"=G{r_cum_x}")]:
        c = w2.cell(row=r, column=col_i, value=f); c.font = BOLD; c.number_format = PROFIT
profit_color(w2, f"B{G+2}:E{G+3}")
w2.cell(row=G + 5, column=1, value=("Read: the pilot is an investment (red) in both scenarios. At the NEED line the machine reaches break-even from M7 "
    "(profit ≈ $0/mo — that is the definition of the line). At the OPTIMIZE line the same spend returns ~+$6K/mo from M7 to reinvest. Green = profit, red = loss.")).font = SUB
w2.merge_cells(start_row=G + 5, start_column=1, end_row=G + 5, end_column=8)

# after-M4 states
A = end_opt + 3
w2.cell(row=A, column=1, value="After-M4 states (exit run-rate vs the floor)").font = BOLD
hdrs = ["Retention of ~3,450 pilot FTDs", "Active deps", "GGR / mo", "Net / mo (51.5%)", "vs M7+ run-rate", "Verdict"]
for j, h in enumerate(hdrs):
    c = w2.cell(row=A + 1, column=1 + j, value=h); c.fill = HDR_FILL; c.font = HDR_FONT
states = [(0.15, "KILL — bleeds"), (0.25, "Marginal — kill unless trending up"), (0.28, "Break-even — the GO line"), (0.39, "Compounding — scale it")]
for i, (ret, verdict) in enumerate(states):
    r = A + 2 + i
    c = w2.cell(row=r, column=1, value=ret); c.font = BLUE; c.number_format = PCT
    w2.cell(row=r, column=2, value=f"=A{r}*$E${r_cum}").number_format = NUM
    w2.cell(row=r, column=3, value=f"=B{r}*$B$5").number_format = CUR
    w2.cell(row=r, column=4, value=f"=C{r}*$B$8").number_format = CUR
    w2.cell(row=r, column=5, value=f"=D{r}-$B$11").number_format = PROFIT
    w2.cell(row=r, column=6, value=verdict).font = BLACK
    for col in range(2, 6): w2.cell(row=r, column=col).font = BLACK
profit_color(w2, f"E{A+2}:E{A+5}")

F = A + 8
w2.cell(row=F, column=1, value="M4 gate floor (GGR/mo) = M7+ run-rate ÷ net contribution").font = BOLD
c = w2.cell(row=F, column=2, value="=B11/B8"); c.font = BLACK; c.number_format = CUR

w2.column_dimensions["A"].width = 44
for col in range(2, 9):
    w2.column_dimensions[get_column_letter(col)].width = 15
w2.freeze_panes = "B21"

# ---------------- Sheet 3: P&L Cascade ----------------
MOVING_FILL = PatternFill("solid", start_color="FDE9C8")
FIXED_FILL = PatternFill("solid", start_color="D6F0F7")
w3 = wb.create_sheet("P&L Cascade")
w3["A1"] = "Full P&L Cascade — From GGR to Net Profit (Path C stack)"
w3["A1"].font = H1
w3.merge_cells("A1:F1")
w3["A2"] = ("Read top to bottom: every row takes a real cost out of real GGR. MOVING fees scale with revenue (% of GGR); "
            "FIXED fees are $/month regardless of revenue. Includes the ~5% high-estimate PIX on-ramp fee and treats the Cubeia €5K/mo (M7+) "
            "as additive — deliberately more conservative than the dashboard's at-scale page. "
            "Paths A/B variant: add 9% affiliate → 47.5% net, floor ~$35K.")
w3["A2"].font = SUB
w3.merge_cells("A2:F2")

h3 = ["Line", "Type", "Rate / $ per month", "At the floor (~$28.5K GGR)", "At $100K GGR", "Who gets it / note"]
for j, h in enumerate(h3):
    c = w3.cell(row=4, column=1 + j, value=h); c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = Alignment(horizontal="right" if j in (2, 3, 4) else "left")

GG = 5
w3.cell(row=GG, column=1, value="Gross Gaming Revenue (GGR)").font = BOLD
w3.cell(row=GG, column=2, value="the input").font = SUB
for col, v in [(4, 28500), (5, 100000)]:
    c = w3.cell(row=GG, column=col, value=v); c.font = BLUE; c.number_format = CUR; c.fill = YELLOW
w3.cell(row=GG, column=6, value="stakes − player wins · change the blue cells to test any level").font = SUB

moving = [
    ("− Player bonuses & incentives", 0.25, "welcome / reload / cashback — to players"),
    None,  # NGR subtotal slot
    ("− Game-provider rev share (blended)", 0.105, "slot / live studios"),
    ("− Crypto processing", 0.03, "processors + on-chain fees"),
    ("− PIX on-ramp (HIGH estimate, unconfirmed)", 0.05, "fiat→crypto on-ramp — verify the real fee"),
    ("− Cubeia platform rev share", 0.05, "white-label platform (contract §4)"),
]
r = GG
rate_rows = []
for item in moving:
    r += 1
    if item is None:
        w3.cell(row=r, column=1, value="NGR  (GGR − bonuses)").font = BOLD
        for col in (4, 5):
            gl = get_column_letter(col)
            w3.cell(row=r, column=col, value=f"={gl}{GG}+{gl}{r-1}").number_format = CUR
        w3.cell(row=r, column=6, value="what 'real' revenue actually is").font = SUB
        continue
    label, rate, who = item
    w3.cell(row=r, column=1, value=label).font = BLACK
    tc = w3.cell(row=r, column=2, value="MOVING — % of GGR"); tc.fill = MOVING_FILL; tc.font = Font(name=ARIAL, size=9, bold=True, color="7A4D00")
    rc = w3.cell(row=r, column=3, value=rate); rc.font = BLUE; rc.number_format = PCT
    rate_rows.append(r)
    for col in (4, 5):
        gl = get_column_letter(col)
        w3.cell(row=r, column=col, value=f"=-{gl}${GG}*$C{r}").number_format = CUR
    w3.cell(row=r, column=6, value=who).font = SUB
NC = r + 1
w3.cell(row=NC, column=1, value="NET CONTRIBUTION  (after all moving fees)").font = BOLD
nc_rate = "-".join([f"C{x}" for x in rate_rows])
c = w3.cell(row=NC, column=3, value=f"=1-{nc_rate}"); c.font = BLACK; c.number_format = PCT
for col in (4, 5):
    gl = get_column_letter(col)
    w3.cell(row=NC, column=col, value=f"={gl}{GG}*$C{NC}").number_format = CUR
    w3.cell(row=NC, column=col).font = BOLD
w3.cell(row=NC, column=6, value="51.5% of every GGR dollar survives the moving fees").font = SUB
for col in range(1, 7): w3.cell(row=NC, column=col).fill = TOT_FILL

fixed = [
    ("− Salaries (T + O + G)", 5000, "the team"),
    ("− X VIP Transfer Method", 1143, "staff $952 + tool $191"),
    ("− Cubeia AWS hosting", 1836, "€1,700/mo, Appendix B"),
    ("− Cubeia platform minimum (M7+)", 5400, "€5K/mo from M7 — additive in this model (open question vs 'minimum')"),
    ("− Subscriptions", 1335, "Customer.io 324 + Freshchat 54 + security 281 + Claude 200 + other 476"),
]
fr0 = NC + 1
for i, (label, amt, who) in enumerate(fixed):
    r = fr0 + i
    w3.cell(row=r, column=1, value=label).font = BLACK
    tc = w3.cell(row=r, column=2, value="FIXED — $ / month"); tc.fill = FIXED_FILL; tc.font = Font(name=ARIAL, size=9, bold=True, color="0B5563")
    ac = w3.cell(row=r, column=3, value=amt); ac.font = BLUE; ac.number_format = CUR
    for col in (4, 5):
        w3.cell(row=r, column=col, value=f"=-$C{r}").number_format = CUR
    w3.cell(row=r, column=6, value=who).font = SUB
FX_TOT = fr0 + len(fixed)
w3.cell(row=FX_TOT, column=1, value="Total fixed burn  (M7+ run-rate)").font = BOLD
c = w3.cell(row=FX_TOT, column=3, value=f"=SUM(C{fr0}:C{FX_TOT-1})"); c.font = BLACK; c.number_format = CUR
for col in (4, 5):
    gl = get_column_letter(col)
    w3.cell(row=FX_TOT, column=col, value=f"=SUM({gl}{fr0}:{gl}{FX_TOT-1})").number_format = CUR
NP3 = FX_TOT + 1
w3.cell(row=NP3, column=1, value="NET PROFIT / month").font = BOLD
for col in (4, 5):
    gl = get_column_letter(col)
    c = w3.cell(row=NP3, column=col, value=f"={gl}{NC}+{gl}{FX_TOT}"); c.font = BOLD; c.number_format = PROFIT
w3.cell(row=NP3, column=6, value="the floor column landing at ≈ $0 is the kill line proven line-by-line").font = SUB
for col in range(1, 7): w3.cell(row=NP3, column=col).fill = TOT_FILL
profit_color(w3, f"D{NP3}:E{NP3}")
w3.cell(row=NP3 + 2, column=1, value="Growth marketing: $0 baseline after the pilot — reinvest-all means media is funded out of the Net Profit line, never the envelope.").font = SUB
w3.merge_cells(start_row=NP3 + 2, start_column=1, end_row=NP3 + 2, end_column=6)
w3.cell(row=NP3 + 3, column=1, value="The floor, derived: fixed burn ÷ net-contribution % =").font = BOLD
c = w3.cell(row=NP3 + 3, column=3, value=f"=C{FX_TOT}/C{NC}"); c.font = BLACK; c.number_format = CUR
w3.cell(row=NP3 + 3, column=6, value="exit M4 below this run-rate and flat → KILL").font = SUB
w3.column_dimensions["A"].width = 40
w3.column_dimensions["B"].width = 19
w3.column_dimensions["C"].width = 17
w3.column_dimensions["D"].width = 22
w3.column_dimensions["E"].width = 15
w3.column_dimensions["F"].width = 52

# ---------------- Sheet 4: Fees at Scale ----------------
w4 = wb.create_sheet("Fees at Scale")
w4["A1"] = "Every Fee at Scale — monthly GGR of $100K / $1M / $10M / $100M"
w4["A1"].font = H1
w4.merge_cells("A1:F1")
w4["A2"] = ("MOVING fees scale 1:1 with GGR; FIXED/stepped costs dilute — that is the operating leverage. "
            "Rates link to the P&L Cascade sheet (single source of truth). GoLuna fixed opex is a staffed-for-scale estimate; "
            "$100M column extrapolated (~3% of GGR) — tune to a real hiring plan. Growth marketing 20% of GGR is a steady-state "
            "assumption (media ~17.4% + Anzil 15% mgmt ~2.6%); at $5M+/mo renegotiate or in-house the 15%.")
w4["A2"].font = SUB
w4.merge_cells("A2:F2")
w4["A4"] = "Growth marketing (steady-state, % of GGR)"; w4["A4"].font = BLACK
mk = w4["B4"]; mk.value = 0.20; mk.font = BLUE; mk.number_format = PCT; mk.fill = YELLOW

h4 = ["Line", "Type / rate", "$100K / mo", "$1M / mo", "$10M / mo", "$100M / mo"]
for j, h in enumerate(h4):
    c = w4.cell(row=6, column=1 + j, value=h); c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = Alignment(horizontal="right" if j >= 2 else "left")
G4 = 7
w4.cell(row=G4, column=1, value="Gross Gaming Revenue (GGR)").font = BOLD
for j, v in enumerate([100000, 1000000, 10000000, 100000000]):
    c = w4.cell(row=G4, column=3 + j, value=v); c.font = BLUE; c.number_format = CUR
mv_labels = [
    ("− Player bonuses & incentives (→ players)", rate_rows[0]),
    ("− Game providers, blended (→ studios)", rate_rows[1]),
    ("− Crypto processing (→ processors)", rate_rows[2]),
    ("− PIX on-ramp, high est. (→ on-ramp partner)", rate_rows[3]),
    ("− Cubeia platform 5% (→ Cubeia)", rate_rows[4]),
]
for i, (label, rr) in enumerate(mv_labels):
    r = G4 + 1 + i
    w4.cell(row=r, column=1, value=label).font = BLACK
    tc = w4.cell(row=r, column=2, value=f"MOVING — ='P&L Cascade'!C{rr}"); tc.value = f"='P&L Cascade'!C{rr}"; tc.font = GREEN; tc.number_format = PCT
    for j in range(4):
        gl = get_column_letter(3 + j)
        w4.cell(row=r, column=3 + j, value=f"=-{gl}${G4}*$B{r}").number_format = CUR
MVT = G4 + 6
w4.cell(row=MVT, column=1, value="Total moving fees").font = BOLD
c = w4.cell(row=MVT, column=2, value=f"=SUM(B{G4+1}:B{MVT-1})"); c.font = BLACK; c.number_format = PCT
for j in range(4):
    gl = get_column_letter(3 + j)
    w4.cell(row=MVT, column=3 + j, value=f"=SUM({gl}{G4+1}:{gl}{MVT-1})").number_format = CUR
NC4 = MVT + 1
w4.cell(row=NC4, column=1, value="NET CONTRIBUTION").font = BOLD
c = w4.cell(row=NC4, column=2, value=f"=1-B{MVT}"); c.font = BLACK; c.number_format = PCT
for j in range(4):
    gl = get_column_letter(3 + j)
    cc = w4.cell(row=NC4, column=3 + j, value=f"={gl}{G4}+{gl}{MVT}"); cc.font = BOLD; cc.number_format = CUR
for col in range(1, 7): w4.cell(row=NC4, column=col).fill = TOT_FILL
CMIN = NC4 + 1
w4.cell(row=CMIN, column=1, value="− Cubeia platform minimum, M7+ (FIXED)").font = BLACK
w4.cell(row=CMIN, column=2, value="='P&L Cascade'!C" + str(fr0 + 3)).font = GREEN
w4.cell(row=CMIN, column=2).number_format = CUR
for j in range(4):
    w4.cell(row=CMIN, column=3 + j, value=f"=-$B{CMIN}").number_format = CUR
OPX = CMIN + 1
w4.cell(row=OPX, column=1, value="− GoLuna fixed opex (stepped estimate — staffed for scale)").font = BLACK
w4.cell(row=OPX, column=2, value="FIXED / stepped").font = SUB
for j, v in enumerate([9314, 45000, 330000, 3000000]):
    c = w4.cell(row=OPX, column=3 + j, value=-v); c.font = BLUE; c.number_format = CUR
OP = OPX + 1
w4.cell(row=OP, column=1, value="OPERATING PROFIT  (EBITDA, pre-marketing)").font = BOLD
for j in range(4):
    gl = get_column_letter(3 + j)
    c = w4.cell(row=OP, column=3 + j, value=f"={gl}{NC4}+{gl}{CMIN}+{gl}{OPX}"); c.font = BOLD; c.number_format = PROFIT
profit_color(w4, f"C{OP}:F{OP}")
MK = OP + 1
w4.cell(row=MK, column=1, value="− Growth marketing (media + Anzil 15% mgmt)").font = BLACK
tc = w4.cell(row=MK, column=2, value="=B4"); tc.font = BLACK; tc.number_format = PCT
for j in range(4):
    gl = get_column_letter(3 + j)
    w4.cell(row=MK, column=3 + j, value=f"=-{gl}${G4}*$B$4").number_format = CUR
NP4 = MK + 1
w4.cell(row=NP4, column=1, value="NET PROFIT / month").font = BOLD
for j in range(4):
    gl = get_column_letter(3 + j)
    c = w4.cell(row=NP4, column=3 + j, value=f"={gl}{OP}+{gl}{MK}"); c.font = BOLD; c.number_format = PROFIT
for col in range(1, 7): w4.cell(row=NP4, column=col).fill = TOT_FILL
profit_color(w4, f"C{NP4}:F{NP4}")
MG = NP4 + 1
w4.cell(row=MG, column=1, value="Net margin (% of GGR)").font = BLACK
for j in range(4):
    gl = get_column_letter(3 + j)
    c = w4.cell(row=MG, column=3 + j, value=f"={gl}{NP4}/{gl}{G4}"); c.font = BLACK; c.number_format = PCT
AN = MG + 1
w4.cell(row=AN, column=1, value="Annualised GGR (×12)").font = SUB
for j in range(4):
    gl = get_column_letter(3 + j)
    c = w4.cell(row=AN, column=3 + j, value=f"={gl}{G4}*12"); c.font = SUB; c.number_format = CUR
w4.column_dimensions["A"].width = 46
w4.column_dimensions["B"].width = 16
for col in range(3, 7):
    w4.column_dimensions[get_column_letter(col)].width = 15

wb.save("Anzil_Brazil_Pilot_Forecast.xlsx")
print("saved", TOT, CUM, REM)
